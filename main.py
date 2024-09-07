from fastapi import FastAPI, Depends, HTTPException, status
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import Workbook, load_workbook
from jose import JWTError, jwt
from passlib.context import CryptContext
from datetime import datetime, timedelta
import os

# Secret key for JWT encoding/decoding
SECRET_KEY = "RabinBhattarai"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

app = FastAPI()

# Path to the Excel file
file_path = 'products.xlsx'

# If the Excel file doesn't exist, create it with headers
if not os.path.exists(file_path):
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = 'Products'
    sheet1.append(['Product ID', 'Product Name', 'Quantity', 'RagNo'])
    
    # Create Sheet2 for user credentials
    sheet2 = workbook.create_sheet('Users')
    sheet2.append(['Username', 'Hashed Password'])
    workbook.save(file_path)

# JWT Token-related utilities
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# Pydantic models
class Product(BaseModel):
    id: int
    name: str
    quantity: int
    RagNo: str

class User(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    username: Optional[str] = None

# Utility Functions
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def read_products():
    workbook = load_workbook(file_path)
    sheet = workbook['Products']
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        products.append(Product(id=row[0], name=row[1], quantity=row[2], RagNo=row[3]))
    return products

def save_products(products: List[Product]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Products'
    sheet.append(['Product ID', 'Product Name', 'Quantity', 'RagNo'])
    for product in products:
        sheet.append([product.id, product.name, product.quantity, product.RagNo])
    workbook.save(file_path)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_user_from_excel(username: str):
    workbook = load_workbook(file_path)
    sheet = workbook['Users']
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return {"username": row[0], "hashed_password": row[1]}
    return None

def authenticate_user(username: str, password: str):
    user = get_user_from_excel(username)
    if not user:
        return False
    if not verify_password(password, user['hashed_password']):
        return False
    return user

def get_current_user(token: str = Depends(lambda token: token)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        return TokenData(username=username)
    except JWTError:
        raise credentials_exception

# Routes
@app.post("/token", response_model=Token)
def login_for_access_token(form_data: User):
    user = authenticate_user(form_data.username, form_data.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(data={"sub": user["username"]}, expires_delta=access_token_expires)
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/register/")
def register_user(user: User):
    workbook = load_workbook(file_path)
    sheet = workbook['Users']
    # Check if user already exists
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user.username:
            raise HTTPException(status_code=400, detail="Username already registered")
    hashed_password = get_password_hash(user.password)
    sheet.append([user.username, hashed_password])
    workbook.save(file_path)
    return {"message": "User registered successfully"}

@app.get("/products/", response_model=List[Product])
def get_products(current_user: TokenData = Depends(get_current_user)):
    return read_products()

@app.post("/products/", response_model=Product)
def create_product(product: Product, current_user: TokenData = Depends(get_current_user)):
    products = read_products()
    for existing_product in products:
        if existing_product.id == product.id:
            raise HTTPException(status_code=400, detail="Product ID already exists")
    products.append(product)
    save_products(products)
    return product

@app.put("/products/{product_id}", response_model=Product)
def update_product(product_id: int, updated_product: Product, current_user: TokenData = Depends(get_current_user)):
    products = read_products()
    for index, product in enumerate(products):
        if product.id == product_id:
            products[index] = updated_product
            save_products(products)
            return updated_product
    raise HTTPException(status_code=404, detail="Product not found")

@app.delete("/products/{product_id}")
def delete_product(product_id: int, current_user: TokenData = Depends(get_current_user)):
    products = read_products()
    for product in products:
        if product.id == product_id:
            products.remove(product)
            save_products(products)
            return {"message": "Product deleted successfully"}
    raise HTTPException(status_code=404, detail="Product not found")
